import io
import openpyxl
from django.db import transaction
from django.http import HttpResponse
from rest_framework import status, viewsets
from rest_framework.decorators import action
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from accounts.permissions import IsAdminRole
from .models import (
    Component, CPU, Motherboard, Videocard, Memory,
    CPU_Cooler, Power_Supply, Storage, Configuration, Config_Item,
)
from .serializers import (
    ConfigurationSerializer,
    ConfigurationSaveSerializer,
    CPUSerializer,
    MotherboardSerializer,
    VideocardSerializer,
    MemorySerializer,
    CPUCoolerSerializer,
    PowerSupplySerializer,
    StorageSerializer,
)
from .services import check_configuration_limit
from .techpowerup_client import search_external, _norm as tpu_norm

CATEGORY_MAP = {
    'cpus': (CPU, CPUSerializer),
    'motherboards': (Motherboard, MotherboardSerializer),
    'gpus': (Videocard, VideocardSerializer),
    'videocards': (Videocard, VideocardSerializer),
    'rams': (Memory, MemorySerializer),
    'memory': (Memory, MemorySerializer),
    'coolings': (CPU_Cooler, CPUCoolerSerializer),
    'coolers': (CPU_Cooler, CPUCoolerSerializer),
    'psus': (Power_Supply, PowerSupplySerializer),
    'power-supplies': (Power_Supply, PowerSupplySerializer),
    'storages': (Storage, StorageSerializer),
}

IMPORT_SHEETS = {
    'cpus': 'cpus',
    'motherboards': 'motherboards',
    'gpus': 'gpus',
    'rams': 'rams',
    'coolings': 'coolings',
    'psus': 'psus',
    'storages': 'storages',
}


def _normalize_category(cat):
    aliases = {
        'videocards': 'gpus',
        'memory': 'rams',
        'coolers': 'coolings',
        'power-supplies': 'psus',
    }
    return aliases.get(cat, cat)


def _bool_val(v):
    if v is None or v == '':
        return False
    if isinstance(v, bool):
        return v
    return str(v).strip().lower() in ('1', 'true', 'да', 'yes', 'y')


def _num(v, default=0, float_ok=False):
    if v is None or v == '':
        return default
    try:
        return float(v) if float_ok else int(float(v))
    except (TypeError, ValueError):
        return default


class AdminConfigurationViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated, IsAdminRole]

    def get_queryset(self):
        return (
            Configuration.objects.filter(user=self.request.user)
            .prefetch_related('items__component')
            .order_by('-id')
        )

    def get_serializer_class(self):
        if self.action in ('create', 'update', 'partial_update'):
            return ConfigurationSaveSerializer
        return ConfigurationSerializer

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        instance = serializer.save()
        output = ConfigurationSerializer(instance, context=self.get_serializer_context())
        return Response(output.data, status=status.HTTP_201_CREATED)

    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        output = ConfigurationSerializer(instance, context=self.get_serializer_context())
        return Response(output.data)


class AdminComponentView(APIView):
    permission_classes = [IsAuthenticated, IsAdminRole]

    def get(self, request):
        category = _normalize_category(request.query_params.get('category', 'cpus'))
        model_cls, serializer_cls = CATEGORY_MAP.get(category, (None, None))
        if not model_cls:
            return Response({'detail': 'Неизвестная категория.'}, status=400)
        qs = model_cls.objects.all().order_by('id')[:500]
        return Response(serializer_cls(qs, many=True).data)

    def post(self, request):
        category = _normalize_category(request.data.get('category'))
        model_cls, serializer_cls = CATEGORY_MAP.get(category, (None, None))
        if not serializer_cls:
            return Response({'detail': 'Неизвестная категория.'}, status=400)
        data = {**request.data}
        data['category'] = category if category in ('gpus', 'rams', 'coolings', 'psus') else request.data.get('category', category)
        if category == 'gpus':
            data['category'] = 'gpus'
        elif category == 'rams':
            data['category'] = 'rams'
        elif category == 'coolings':
            data['category'] = 'coolings'
        elif category == 'psus':
            data['category'] = 'psus'
        else:
            data.setdefault('category', category)

        serializer = serializer_cls(data=data)
        serializer.is_valid(raise_exception=True)
        obj = serializer.save()
        return Response(serializer_cls(obj).data, status=status.HTTP_201_CREATED)


def _component_instance_and_serializer(pk):
    try:
        comp = Component.objects.get(pk=pk)
    except Component.DoesNotExist:
        return None, None, None
    cat = _normalize_category(comp.category)
    pair = CATEGORY_MAP.get(cat)
    if not pair:
        return comp, None, cat
    model_cls, serializer_cls = pair
    try:
        obj = model_cls.objects.get(pk=pk)
    except model_cls.DoesNotExist:
        return comp, None, cat
    return obj, serializer_cls, cat


class AdminComponentDetailView(APIView):
    permission_classes = [IsAuthenticated, IsAdminRole]

    def get(self, request, pk):
        obj, serializer_cls, cat = _component_instance_and_serializer(pk)
        if not obj:
            return Response({'detail': 'Компонент не найден.'}, status=404)
        if not serializer_cls:
            return Response({'detail': 'Неизвестная категория компонента.'}, status=400)
        return Response(serializer_cls(obj).data)

    def put(self, request, pk):
        obj, serializer_cls, cat = _component_instance_and_serializer(pk)
        if not obj:
            return Response({'detail': 'Компонент не найден.'}, status=404)
        if not serializer_cls:
            return Response({'detail': 'Неизвестная категория компонента.'}, status=400)
        data = {**request.data, 'category': cat if cat in ('gpus', 'rams', 'coolings', 'psus') else request.data.get('category', cat)}
        serializer = serializer_cls(obj, data=data, partial=False)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer_cls(obj).data)

    def patch(self, request, pk):
        obj, serializer_cls, cat = _component_instance_and_serializer(pk)
        if not obj:
            return Response({'detail': 'Компонент не найден.'}, status=404)
        if not serializer_cls:
            return Response({'detail': 'Неизвестная категория компонента.'}, status=400)
        data = dict(request.data)
        if 'category' not in data:
            data['category'] = obj.category
        serializer = serializer_cls(obj, data=data, partial=True)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer_cls(obj).data)

    def delete(self, request, pk):
        try:
            comp = Component.objects.get(pk=pk)
        except Component.DoesNotExist:
            return Response({'detail': 'Компонент не найден.'}, status=404)
        comp.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


class AdminImportXlsxView(APIView):
    permission_classes = [IsAuthenticated, IsAdminRole]
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request):
        upload = request.FILES.get('file')
        if not upload:
            return Response({'detail': 'Прикрепите файл .xlsx.'}, status=400)
        if not upload.name.lower().endswith('.xlsx'):
            return Response({'detail': 'Допустим только формат .xlsx.'}, status=400)

        try:
            wb = openpyxl.load_workbook(upload, read_only=True, data_only=True)
        except Exception as exc:
            return Response({'detail': f'Не удалось прочитать файл: {exc}'}, status=400)

        stats = {}
        errors = []

        with transaction.atomic():
            for sheet_name, cat_key in IMPORT_SHEETS.items():
                if sheet_name not in wb.sheetnames:
                    continue
                ws = wb[sheet_name]
                rows = list(ws.iter_rows(values_only=True))
                if len(rows) < 2:
                    continue
                headers = [str(h).strip().lower() if h else '' for h in rows[0]]
                created = 0
                model_cls, _ = CATEGORY_MAP[cat_key]

                for row_idx, row in enumerate(rows[1:], start=2):
                    if not row or all(c is None or str(c).strip() == '' for c in row):
                        continue
                    row_dict = {}
                    for i, h in enumerate(headers):
                        if h and i < len(row):
                            row_dict[h] = row[i]
                    try:
                        _create_from_row(cat_key, model_cls, row_dict)
                        created += 1
                    except Exception as e:
                        errors.append(f'{sheet_name} строка {row_idx}: {e}')
                stats[sheet_name] = created

        return Response({
            'message': 'Импорт завершён.',
            'imported': stats,
            'errors': errors[:30],
        })


class AdminImportTemplateView(APIView):
    """Скачать пустой шаблон Excel для импорта комплектующих."""
    permission_classes = [IsAuthenticated, IsAdminRole]

    def get(self, request):
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        sheet_columns = {
            'cpus': [
                'manufacturer', 'name', 'socket', 'cores', 'threads',
                'base_clock_ghz', 'boost_clock_ghz', 'l3_cache_mb', 'tdp',
                'integrated_graphics', 'memory_channels',
            ],
            'motherboards': [
                'manufacturer', 'name', 'socket', 'chipset', 'form_factor',
                'ram_slots', 'ram_max', 'ram_type', 'ram_speed', 'pcie_ver',
                'pcie_slots', 'm2_slots', 'm2_type', 'sata_ports', 'wifi', 'bluetooth',
            ],
            'gpus': [
                'manufacturer', 'name', 'gpu_chip', 'chipset', 'capacity', 'pcie',
                'form_factor', 'tdp', 'power_connectors', 'length', 'outputs', 'boost_clock',
            ],
            'rams': ['manufacturer', 'name', 'type', 'capacity', 'speed_mhz', 'timings', 'voltage'],
            'coolings': ['manufacturer', 'name', 'height', 'fan_speed', 'noise_level', 'tdp', 'socket'],
            'psus': [
                'manufacturer', 'name', 'wattage', 'efficiency_rating',
                'connectors_24pin', 'connectors_cpu4_4pin', 'connectors_cpu_8pin',
                'connectors_pcie_6_2pin', 'connectors_pcie_8pin', 'connectors_pcie_12pin', 'connectors_sata',
            ],
            'storages': [
                'manufacturer', 'name', 'type', 'form_factor', 'capacity',
                'interface', 'read_speed', 'write_speed',
            ],
        }
        for sheet_name, cols in sheet_columns.items():
            ws = wb.create_sheet(sheet_name)
            ws.append(cols)
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        response = HttpResponse(
            buf.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="components_import_template.xlsx"'
        return response


def _filter_not_in_catalog(results, kind):
    """Оставить только позиции, которых ещё нет в нашей БД."""
    cat_key = 'gpus'
    model_cls, _ = CATEGORY_MAP[cat_key]
    existing = {tpu_norm(n) for n in model_cls.objects.values_list('name', flat=True)}
    filtered = []
    for row in results:
        name = (row.get('name') or '').strip()
        if not name:
            continue
        if tpu_norm(name) in existing:
            continue
        if model_cls.objects.filter(name__iexact=name).exists():
            continue
        filtered.append(row)
    return filtered


class AdminTechPowerUpSearchView(APIView):
    permission_classes = [IsAuthenticated, IsAdminRole]

    def get(self, request):
        q = (request.query_params.get('q') or '').strip()
        kind = (request.query_params.get('type') or 'gpu').lower()
        if len(q) < 2:
            return Response({'detail': 'Введите минимум 2 символа для поиска.'}, status=400)
        try:
            raw = search_external(kind, q, limit=int(request.query_params.get('limit', 30)))
            results = _filter_not_in_catalog(raw, kind)
        except ValueError as exc:
            return Response({'detail': str(exc)}, status=400)
        except Exception as exc:
            return Response({'detail': f'Не удалось загрузить каталог: {exc}'}, status=502)
        labels = {'gpu': 'видеокарт'}
        skipped = len(raw) - len(results)
        msg = f'Новых позиций ({labels.get(kind, kind)}): {len(results)}'
        if skipped:
            msg += f' (в каталоге уже есть: {skipped})'
        if not results and raw:
            msg = 'Все найденные позиции уже есть в каталоге. Уточните запрос.'
        return Response({
            'query': q,
            'type': kind,
            'source': 'open_gpu_catalog',
            'results': results,
            'message': msg,
        })


class AdminTechPowerUpImportView(APIView):
    """Сохранить отредактированный черновик в каталог."""
    permission_classes = [IsAuthenticated, IsAdminRole]

    def post(self, request):
        data = request.data
        name = (data.get('name') or '').strip()
        manufacturer = (data.get('manufacturer') or '').strip()
        if not name or not manufacturer:
            return Response({'detail': 'Поля «производитель» и «название» обязательны.'}, status=400)

        kind = (data.get('import_kind') or data.get('category') or 'gpus').lower()
        cat_key = _normalize_category(kind)
        if cat_key not in CATEGORY_MAP:
            return Response({'detail': 'Неподдерживаемый тип компонента для импорта.'}, status=400)

        model_cls, serializer_cls = CATEGORY_MAP[cat_key]
        row = {**data, 'manufacturer': manufacturer, 'name': name}
        try:
            _create_from_row(cat_key, model_cls, row)
        except Exception as exc:
            return Response({'detail': f'Ошибка сохранения: {exc}'}, status=400)

        obj = model_cls.objects.filter(name=name, manufacturer=manufacturer).order_by('-id').first()
        labels = {
            'gpus': 'Видеокарта',
            'cpus': 'Процессор',
            'storages': 'Накопитель',
        }
        return Response({
            'message': f'{labels.get(cat_key, "Компонент")} добавлен в каталог.',
            'component': serializer_cls(obj).data if obj else None,
        }, status=201)


def _create_from_row(cat_key, model_cls, d):
    mfr = d.get('manufacturer') or d.get('производитель') or 'Unknown'
    name = d.get('name') or d.get('название') or 'Без названия'
    base = {'category': cat_key, 'manufacturer': str(mfr), 'name': str(name)}

    if cat_key == 'cpus':
        CPU.objects.create(
            **base,
            socket=str(d.get('socket', '')),
            cores=_num(d.get('cores'), 4),
            threads=_num(d.get('threads'), 8),
            base_clock_ghz=_num(d.get('base_clock_ghz'), 3.0, float_ok=True),
            boost_clock_ghz=_num(d.get('boost_clock_ghz'), 4.0, float_ok=True),
            l3_cache_mb=_num(d.get('l3_cache_mb'), 16),
            tdp=_num(d.get('tdp'), 65),
            integrated_graphics=_bool_val(d.get('integrated_graphics')),
            memory_channels=_num(d.get('memory_channels'), 2),
        )
    elif cat_key == 'motherboards':
        Motherboard.objects.create(
            **base,
            socket=str(d.get('socket', '')),
            chipset=str(d.get('chipset', '')),
            form_factor=str(d.get('form_factor', 'ATX')),
            ram_slots=_num(d.get('ram_slots'), 4),
            ram_max=_num(d.get('ram_max'), 128),
            ram_type=str(d.get('ram_type', 'DDR4')),
            ram_speed=_num(d.get('ram_speed'), 3200),
            pcie_ver=str(d.get('pcie_ver', 'PCI-e 4.0')),
            pcie_slots=_num(d.get('pcie_slots'), 2),
            m2_slots=_num(d.get('m2_slots'), 2),
            m2_type=str(d.get('m2_type', 'M.2 NVMe')),
            sata_ports=_num(d.get('sata_ports'), 4),
            wifi=_bool_val(d.get('wifi')),
            bluetooth=_bool_val(d.get('bluetooth')),
        )
    elif cat_key == 'gpus':
        Videocard.objects.create(
            **base,
            gpu_chip=str(d.get('gpu_chip', '')),
            chipset=str(d.get('chipset', '')),
            capacity=_num(d.get('capacity'), 8),
            pcie=str(d.get('pcie', 'PCI-e 4.0')),
            form_factor=str(d.get('form_factor', 'ATX')),
            tdp=_num(d.get('tdp'), 200),
            power_connectors=str(d.get('power_connectors', '')),
            length=_num(d.get('length'), 300),
            outputs=str(d.get('outputs', '')),
            boost_clock=_num(d.get('boost_clock'), 2000),
        )
    elif cat_key == 'rams':
        Memory.objects.create(
            **base,
            type=str(d.get('type', 'DDR4')),
            capacity=_num(d.get('capacity'), 16),
            speed_mhz=_num(d.get('speed_mhz'), 3200),
            timings=str(d.get('timings', '')),
            voltage=_num(d.get('voltage'), 1.35, float_ok=True),
        )
    elif cat_key == 'coolings':
        CPU_Cooler.objects.create(
            **base,
            height=_num(d.get('height'), 150),
            fan_speed=str(d.get('fan_speed', '')),
            noise_level=_num(d.get('noise_level'), 30, float_ok=True),
            tdp=_num(d.get('tdp'), 150),
            socket=str(d.get('socket', '')),
        )
    elif cat_key == 'psus':
        Power_Supply.objects.create(
            **base,
            wattage=_num(d.get('wattage'), 650),
            efficiency_rating=str(d.get('efficiency_rating', '80+ Bronze')),
            connectors_24pin=_num(d.get('connectors_24pin'), 1),
            connectors_cpu4_4pin=_num(d.get('connectors_cpu4_4pin'), 1),
            connectors_cpu_8pin=_num(d.get('connectors_cpu_8pin'), 0),
            connectors_pcie_6_2pin=_num(d.get('connectors_pcie_6_2pin'), 2),
            connectors_pcie_8pin=_num(d.get('connectors_pcie_8pin'), 0),
            connectors_pcie_12pin=_num(d.get('connectors_pcie_12pin'), 0),
            connectors_sata=_num(d.get('connectors_sata'), 4),
        )
    elif cat_key == 'storages':
        Storage.objects.create(
            **base,
            type=str(d.get('type', 'SSD')),
            form_factor=str(d.get('form_factor', 'M.2')),
            capacity=_num(d.get('capacity'), 1000),
            interface=str(d.get('interface', 'NVMe')),
            read_speed=_num(d.get('read_speed'), 3000),
            write_speed=_num(d.get('write_speed'), 2000),
        )
